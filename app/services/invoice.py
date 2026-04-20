from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import os
from .pdf_conversion import excel2pdf, autofit_columns

# router = APIRouter()

# @router.post("/invoice")
def generate_invoice_file(payload: dict, invoice_number: str) -> str:
    
    # Generate invoice PDF
    os.makedirs("generated_invoices", exist_ok=True)
    excel_path = f"generated_invoices/{invoice_number}.xlsx"
    pdf_path = f"generated_invoices/{invoice_number}.pdf"

    # Default font style
    font = Font(bold=True, color="0000FF", name="Arial")

    header_fill = PatternFill(start_color="3960FA", end_color="3960FA", fill_type="solid")
    tab_fill1 = PatternFill(start_color="C7D1F7", end_color="C7D1F7", fill_type="solid")
    tab_fill2 = PatternFill(start_color="93A2DC", end_color="93A2DC", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "City Light Invoice"

    ws["A1"] = "Invoice"
    ws["A1"].font = Font(size=16, bold=True, color="0000FF")
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Date: {payload['date_created']}"
    ws["A3"] = f"Invoice Number: {invoice_number}"

    ws["A4"] = "Customer Details"
    ws["A4"].font = font
    ws["A5"] = payload["client_name"]
    ws["A6"] = payload["client_address"]
    ws["A7"] = payload["client_number"]

    ws.merge_cells("A6:D6")
    ws.merge_cells("A7:D7")
    ws.merge_cells("A8:D8")
    ws.merge_cells("A9:D9")

    headers = [
        "#",
        "Item Description",
        "Quantity",
        "Price Ex.\nVat",
        "Price Inc.\nVat",
        "Total Price"
    ]

    wrap_align = Alignment(wrap_text=True, horizontal="center", vertical="center")

    ws["A11"].value = headers[0]
    ws["B11"].value = headers[1]
    ws.merge_cells("B11:D11")
    ws["F11"].value = headers[2]
    ws["G11"].value = headers[3]
    ws["H11"].value = headers[4]
    ws["I11"].value = headers[5]

    for col in ["A", "B", "F", "G", "H", "I"]:
        ws[f"{col}11"].alignment = wrap_align

    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.row_dimensions[11].height = 26

    header_row = 11
    start_row = header_row + 1

    for i, item in enumerate(payload["items"], start=1):
        row = start_row + i - 1

        unit_price_ex = item["unit_price_ex_vat"]
        unit_price_inc = item["unit_price_inc_vat"]
        line_total = item["line_total"]
        description = item["description"].lower()

        # A: Row number
        ws.cell(row=row, column=1, value=i)

        # B-E: Item Description
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=item["description"])

        # F: Quantity
        if description == "tax total":
            ws.cell(row=row, column=6, value="")
        else:
            ws.cell(row=row, column=6, value=item["quantity"])

        # G: Net Unit Price
        if description == "labour":
            ws.cell(row=row, column=7, value="")
        else:
            ws.cell(row=row, column=7, value=unit_price_ex)

        # H: Unit Price Including VAT
        if description == "labour":
            ws.cell(row=row, column=8, value=unit_price_ex)
        else:
            ws.cell(row=row, column=8, value=unit_price_inc)

        # I: Total Price
        ws.cell(row=row, column=9, value=line_total)

    current_row = start_row + len(payload["items"])

    grand_total_label = ws.cell(row=current_row + 1, column=8, value="Grand Total")
    grand_total_label.font = Font(bold=True)
    grand_total_label.alignment = Alignment(horizontal="right")

    total_column_letter = get_column_letter(9)
    grand_total_formula = f"=SUM({total_column_letter}{start_row}:{total_column_letter}{current_row})"
    totals = ws.cell(row=current_row + 1, column=9, value=grand_total_formula)
    totals.font = Font(bold=True)
    totals.number_format = '"R"#,##0.00'

    for col in range(1, 10):
        ws.cell(11, col).font = Font(bold=True)
        ws.cell(11, col).alignment = Alignment(horizontal="center")
        ws.cell(11, col).fill = header_fill

    # Fill table rows color
    for row in range(start_row, current_row + 1, 2):
        for col in range(1, 10):
            ws.cell(row, col).fill = tab_fill1

    for row in range(start_row + 1, current_row + 1, 2):
        for col in range(1, 10):
            ws.cell(row, col).fill = tab_fill2

    # Autofit
    autofit_columns(
        ws,
        start_row=11,
        end_row=current_row + 2,
        start_col=1,
        end_col=1
    )
    autofit_columns(
        ws,
        start_row=11,
        end_row=current_row + 2,
        start_col=9,
        end_col=9
    )

    # Banking details
    ws[f"A{current_row + 3}"] = "Banking Details"
    ws[f"A{current_row + 3}"].font = font
    ws[f"A{current_row + 4}"] = "Bank: Bidvest Bank Alliance"
    ws[f"A{current_row + 5}"] = "Branch Code: 683000"
    ws[f"A{current_row + 6}"] = "Account Holder: Leeroy Antony Muzondi"
    ws[f"A{current_row + 7}"] = "Account Number: 7860 2801 824"
    ws[f"A{current_row + 8}"] = "Account Type: Current"
    ws[f"A{current_row + 10}"] = "Note: Make sure the bank is Bidvest Bank Alliance and the Branch Code is 683000."

    ws.merge_cells(f"A{current_row + 3}:H{current_row + 3}")
    ws.merge_cells(f"A{current_row + 4}:H{current_row + 4}")
    ws.merge_cells(f"A{current_row + 5}:H{current_row + 5}")
    ws.merge_cells(f"A{current_row + 6}:H{current_row + 6}")
    ws.merge_cells(f"A{current_row + 7}:H{current_row + 7}")
    ws.merge_cells(f"A{current_row + 8}:H{current_row + 8}")
    ws.merge_cells(f"A{current_row + 10}:H{current_row + 10}")

    ws[f"A{current_row + 10}"].font = Font(bold=False, italic=True)
    ws[f"A{current_row + 10}"].alignment = Alignment(wrap_text=True, horizontal="center")

    wb.save(excel_path)
    excel2pdf(excel_path)

    return pdf_path