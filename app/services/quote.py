from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import datetime
import os 
from app.models import Quotes
from .pdf_conversion import excel2pdf, autofit_columns
from app.schemas import QuoteRequest
from fastapi import APIRouter, Depends
import re
from sqlalchemy.orm import Session
from app.database import get_db

router = APIRouter()

def slugify_name(name: str):
    name = name.lower()
    name = re.sub(r'[^a-z0-9]', '', name)
    return name

# generate

# @router.post("/quote")
def qoute(data: QuoteRequest, db: Session = Depends(get_db)):
    
    new_qoute = Quotes(
        client_name=data.client_name,
        client_address=data.client_address,
        client_date=datetime.date.today(),
    )
    db.add(new_qoute)
    db.commit()
    db.refresh(new_qoute)

    # generate global sequential quote number
    slug = slugify_name(data.client_name)
    sequence = f"{new_qoute.id:04d}"
    qoute_number = f"{slug}-{sequence}"

    # generate quote pdf
    os.makedirs("generated_quotes", exist_ok=True)
    pdf_path = f"generated_quotes/{qoute_number}.pdf"


    # Defalut font style
    font = Font(bold=True, color="0000FF", name="Arial")

    # Block fill
    fill = PatternFill(start_color="FF98AFC7", end_color="FF98AFC7", fill_type="solid")

    header_fill = PatternFill(start_color="3960FA", end_color="3960FA", fill_type="solid")
    tab_fill1 = PatternFill(start_color="C7D1F7", end_color="C7D1F7", fill_type="solid")
    tab_fill2 = PatternFill(start_color="93A2DC", end_color="93A2DC", fill_type="solid")


    wb = Workbook()
    ws = wb.active
    ws.title = "City Light Quote"
    ws["A1"] = "Quotation"
    ws["A1"].font = Font(size=16, bold=True, color="0000FF")
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Quote Number: {qoute_number}"
    ws["A3"] = f"Date: {datetime.date.today().strftime('%Y-%m-%d')}"
    ws["A4"] = f"Valid Until: {(datetime.date.today() + datetime.timedelta(days=14)).strftime('%Y-%m-%d')}"

    ws["A6"] = "Quotation From"
    ws["A6"].font = font
    ws["A7"] = "Lee Electrician"
    ws ["A8"] = "South Africa"
    ws["A9"] = "+27 788 467 030"

    ws.merge_cells("A6:D6")
    ws.merge_cells("A7:D7")
    ws.merge_cells("A8:D8")
    ws.merge_cells("A9:D9")

    ws["E6"] = "Quotation To"
    ws["E6"].font = font 
    ws["E7"] = data.client_name
    ws["E8"] = data.client_address
    ws["E9"] = data.client_city

    ws.merge_cells("E6:H6")
    ws.merge_cells("E7:H7")
    ws.merge_cells("E8:H8")
    ws.merge_cells("E9:H9")

    headers = ["#", "Item Description", "Quantity", "Unit Price", "Total Price"]

    ws["A11"] = headers[0]
    ws["B11"] = headers[1]
    ws.merge_cells("B11:E11")
    ws["F11"] = headers[2]
    ws["G11"] = headers[3]
    ws["H11"] = headers[4]

    header_row = 11
    start_row =  header_row +1 

    for i, item in enumerate(data.items.values(), start=1):
        row =  start_row + i - 1
        unit_price = item.unit_price * 1.2
        total = item.quantity * unit_price
        

        # A: Row number
        ws.cell(row=row, column=1, value= i)

        # B-D: Item Description
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=item.description)

        # F: Quantity
        ws.cell(row=row, column=6, value=item.quantity)

        # G: Unit Price
        ws.cell(row=row, column=7, value=unit_price)

        # H: Total Price 
        ws.cell(row=row, column=8, value=total)

        
    # Add data rows
    start_row = 11 
    current_row = start_row + len(data.items) + 1

    for row in range(start_row, current_row):
        for col in range(2, 4):
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

    grand_total_cell = ws.cell(row=current_row, column=7, value="Grand Total")
    grand_total_cell.font = Font(bold=True)
    grand_total_cell.alignment = Alignment(horizontal="right")

    total_coumn_letter = get_column_letter(8)
    grand_total_formula = (
    f"=ROUND(SUM({total_coumn_letter}{start_row}:{total_coumn_letter}{current_row - 1}), 2)"
    )

    totals = ws.cell(row=current_row, column=8, value=grand_total_formula)
    totals.font = Font(bold=True)
    totals.number_format = '"R"#,##0.00'


    for col in range(1, 9):
        ws.cell(11, col).font = Font(bold=True)
        ws.cell(11, col).alignment = Alignment(horizontal="center")
        ws.cell(11, col).fill = header_fill


    # fill table rows color
    for row in range(start_row + 1, current_row, 2):
        for col in range(1, 9):
            ws.cell(row, col).fill = tab_fill1
    for row in range(start_row + 2, current_row, 2):
        for col in range(1, 9):
            ws.cell(row, col).fill = tab_fill2
        
    #ws.add_table(tab)
    # Autofit only table area
    autofit_columns(
        ws,
        start_row=11,
        end_row=current_row + 1,
        start_col=1,
        end_col=1
    )
    autofit_columns(
        ws,
        start_row=11,
        end_row=current_row + 1,
        start_col=8,
        end_col=8
    )
    # banking details
    ws["A" + str(current_row + 2)] = "Banking Details"
    ws.merge_cells(f"A{current_row + 2}:H{current_row + 2}")
    ws["A" + str(current_row + 2)].font = font
    ws["A" + str(current_row + 3)] = "Bank: Bidvest Bank Alliance"
    ws.merge_cells(f"A{current_row + 3}:H{current_row + 3}")
    ws["A" + str(current_row + 4)] = "Branch Code: 683000"
    ws.merge_cells(f"A{current_row + 4}:H{current_row + 4}")
    ws["A" + str(current_row + 5)] = "Account Holder: Leeroy Antony Muzondi"
    ws.merge_cells(f"A{current_row + 5}:H{current_row + 5}")
    ws["A" + str(current_row + 6)] = "Account Number: 7860 2801 824"
    ws.merge_cells(f"A{current_row + 6}:H{current_row + 6}")
    ws["A" + str(current_row + 7)] = "Account Type: Current"
    ws.merge_cells(f"A{current_row + 7}:H{current_row + 7}")
    ws["A" + str(current_row + 9)] = "Note: Make sure the bank is BidvestBank Alliance and the Branch Code is 683000."
    ws.merge_cells(f"A{current_row + 9}:H{current_row + 9}")

    ws["A" + str(current_row + 2)].font = font
    ws["A" + str(current_row + 9)].font = Font(bold=False, italic=True)
    ws["A" + str(current_row + 9)].alignment = Alignment(wrap_text=True,horizontal="center")

    # Terms and conditions
    grand_total = 0

    for r in range(start_row, current_row):
        value = ws.cell(row=r, column=8).value
        
        if isinstance(value, (int, float)):
            grand_total += value

    # write numeric total (for backend logic)
    totals = ws.cell(row=current_row, column=8, value=grand_total)

    # now this works
    calculated_total = grand_total
    part_payment = 0.7 * calculated_total

    part_payment = 0.7 * calculated_total
    ws['A' + str(current_row + 11)] = "Terms and Conditions:"
    ws.merge_cells(f"A{current_row + 11}:H{current_row + 11}")
    ws['A' + str(current_row + 12)] = "1. All prices are Inclusive of VAT."
    ws.merge_cells(f"A{current_row + 12}:H{current_row + 12}")
    ws['A' + str(current_row + 13)] = f"2. 70 % Payment of R{part_payment:.2f} is must be made before work commences."
    ws.merge_cells(f"A{current_row + 13}:H{current_row + 13}")

    excel_path = f"generated_quotes/{qoute_number}.xlsx"
    wb.save(excel_path)
    pdf = excel2pdf(excel_path)
    pdf_path = f"generated_quotes/{qoute_number}.pdf"
    new_qoute.qoute_number = qoute_number
    new_qoute.client_qoute_pdf = pdf_path
    db.commit()
    db.refresh(new_qoute)
    return FileResponse(
        path=pdf_path,
        media_type='application/pdf',
        filename=f"{qoute_number} - quote.pdf"
    )

def generate_quote_file(data: QuoteRequest, qoute_number: str):
    
    os.makedirs("generated_quotes", exist_ok=True)
    excel_path = f"generated_quotes/{qoute_number}.xlsx"
    pdf_path = f"generated_quotes/{qoute_number}.pdf"

    # Defalut font style
    font = Font(bold=True, color="0000FF", name="Arial")

    # Block fill
    fill = PatternFill(start_color="FF98AFC7", end_color="FF98AFC7", fill_type="solid")

    header_fill = PatternFill(start_color="3960FA", end_color="3960FA", fill_type="solid")
    tab_fill1 = PatternFill(start_color="C7D1F7", end_color="C7D1F7", fill_type="solid")
    tab_fill2 = PatternFill(start_color="93A2DC", end_color="93A2DC", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "City Light Quote"

    ws["A1"] = "Quotation"
    ws["A1"].font = Font(size=16, bold=True, color="0000FF")
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Quote Number: {qoute_number}"
    ws["A3"] = f"Date: {datetime.date.today().strftime('%Y-%m-%d')}"

    ws["A6"] = "Quotation From"
    ws["A6"].font = font
    ws["A7"] = "Lee Electrician"
    ws ["A8"] = "South Africa"
    ws["A9"] = "+27 788 467 030"

    ws.merge_cells("A6:D6")
    ws.merge_cells("A7:D7")
    ws.merge_cells("A8:D8")
    ws.merge_cells("A9:D9")

    ws["E6"] = "Quotation To"
    ws["E6"].font = font 
    ws["E7"] = data.client_name
    ws["E8"] = data.client_address
    ws["E9"] = data.client_city

    ws.merge_cells("E6:H6")
    ws.merge_cells("E7:H7")
    ws.merge_cells("E8:H8")
    ws.merge_cells("E9:H9")

    headers = ["#", "Item Description", "Quantity", "Unit Price", "Total Price"]

    ws["A11"] = headers[0]
    ws["B11"] = headers[1]
    ws.merge_cells("B11:E11")
    ws["F11"] = headers[2]
    ws["G11"] = headers[3]
    ws["H11"] = headers[4]

    header_row = 11
    start_row =  header_row +1 

    for i, item in enumerate(data.items.values(), start=1):
        row =  start_row + i - 1
        unit_price = item.unit_price * 1.2
        total = item.quantity * unit_price
        

        # A: Row number
        ws.cell(row=row, column=1, value= i)

        # B-D: Item Description
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=item.description)

        # F: Quantity
        ws.cell(row=row, column=6, value=item.quantity)

        # G: Unit Price
        ws.cell(row=row, column=7, value=unit_price)

        # H: Total Price 
        ws.cell(row=row, column=8, value=total)

        
    # Add data rows
    start_row = 11 
    current_row = start_row + len(data.items) + 1

    for row in range(start_row, current_row):
        for col in range(2, 4):
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

    grand_total_cell = ws.cell(row=current_row, column=7, value="Grand Total")
    grand_total_cell.font = Font(bold=True)
    grand_total_cell.alignment = Alignment(horizontal="right")

    total_coumn_letter = get_column_letter(8)
    grand_total_formula = (
    f"=ROUND(SUM({total_coumn_letter}{start_row}:{total_coumn_letter}{current_row - 1}), 2)"
    )

    totals = ws.cell(row=current_row, column=8, value=grand_total_formula)
    totals.font = Font(bold=True)
    totals.number_format = '"R"#,##0.00'


    for col in range(1, 9):
        ws.cell(11, col).font = Font(bold=True)
        ws.cell(11, col).alignment = Alignment(horizontal="center")
        ws.cell(11, col).fill = header_fill


    # fill table rows color
    for row in range(start_row + 1, current_row, 2):
        for col in range(1, 9):
            ws.cell(row, col).fill = tab_fill1
    for row in range(start_row + 2, current_row, 2):
        for col in range(1, 9):
            ws.cell(row, col).fill = tab_fill2
        
    #ws.add_table(tab)
    # Autofit only table area
    autofit_columns(
        ws,
        start_row=11,
        end_row=current_row + 1,
        start_col=1,
        end_col=1
    )
    autofit_columns(
        ws,
        start_row=11,
        end_row=current_row + 1,
        start_col=8,
        end_col=8
    )
    # banking details
    ws["A" + str(current_row + 2)] = "Banking Details"
    ws.merge_cells(f"A{current_row + 2}:H{current_row + 2}")
    ws["A" + str(current_row + 2)].font = font
    ws["A" + str(current_row + 3)] = "Bank: Bidvest Bank Alliance"
    ws.merge_cells(f"A{current_row + 3}:H{current_row + 3}")
    ws["A" + str(current_row + 4)] = "Branch Code: 683000"
    ws.merge_cells(f"A{current_row + 4}:H{current_row + 4}")
    ws["A" + str(current_row + 5)] = "Account Holder: Leeroy Antony Muzondi"
    ws.merge_cells(f"A{current_row + 5}:H{current_row + 5}")
    ws["A" + str(current_row + 6)] = "Account Number: 7860 2801 824"
    ws.merge_cells(f"A{current_row + 6}:H{current_row + 6}")
    ws["A" + str(current_row + 7)] = "Account Type: Current"
    ws.merge_cells(f"A{current_row + 7}:H{current_row + 7}")
    ws["A" + str(current_row + 9)] = "Note: Make sure the bank is BidvestBank Alliance and the Branch Code is 683000."
    ws.merge_cells(f"A{current_row + 9}:H{current_row + 9}")

    ws["A" + str(current_row + 2)].font = font
    ws["A" + str(current_row + 9)].font = Font(bold=False, italic=True)
    ws["A" + str(current_row + 9)].alignment = Alignment(wrap_text=True,horizontal="center")

    # Terms and conditions
    grand_total = 0

    for r in range(start_row, current_row):
        value = ws.cell(row=r, column=8).value
        
        if isinstance(value, (int, float)):
            grand_total += value

    # write numeric total (for backend logic)
    totals = ws.cell(row=current_row, column=8, value=grand_total)

    # now this works
    calculated_total = grand_total
    part_payment = 0.7 * calculated_total

    part_payment = 0.7 * calculated_total
    ws['A' + str(current_row + 11)] = "Terms and Conditions:"
    ws.merge_cells(f"A{current_row + 11}:H{current_row + 11}")
    ws['A' + str(current_row + 12)] = "1. All prices are Inclusive of VAT."
    ws.merge_cells(f"A{current_row + 12}:H{current_row + 12}")
    ws['A' + str(current_row + 13)] = f"2. 70 % Payment of R{part_payment:.2f} is must be made before work commences."
    ws.merge_cells(f"A{current_row + 13}:H{current_row + 13}")

    excel_path = f"generated_quotes/{qoute_number}.xlsx"

    wb.save(excel_path)

    excel2pdf(excel_path)

    return pdf_path

@router.post("/quote/preview")
def preview_quote(data: QuoteRequest):

    slug = slugify_name(data.client_name)

    preview_number = f"{slug}-preview"

    pdf_path = generate_quote_file(data, preview_number)

    return FileResponse(
        pdf_path,
        media_type="application/pdf",
        filename="quote-preview.pdf"
    )

@router.post("/quote/finalize")
def finalize_quote(data: QuoteRequest, db: Session = Depends(get_db)):
    try:
        new_quote = Quotes(
            client_name=data.client_name,
            client_address=data.client_address,
            client_date=datetime.date.today()
        )

        db.add(new_quote)
        db.flush()   

        slug = slugify_name(data.client_name)
        sequence = f"{new_quote.id:04d}"

        quote_number = f"{slug}-{sequence}"

        pdf_path = generate_quote_file(data, quote_number)

        new_quote.client_quote_number = quote_number
        new_quote.client_quote_pdf = pdf_path

        db.commit()
        db.refresh(new_quote)

        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=f"{quote_number}-quote.pdf"
        )
    except Exception as e:
        db.rollback()
        return {"failed to finalize quote": str(e)}