from openpyxl.utils import get_column_letter
from fastapi.responses import FileResponse
from threading import Lock
import win32com.client
import pythoncom
import os

def autofit_columns(ws, start_row, end_row, start_col, end_col):
    for col_idx in range(start_col, end_col + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        if max_length > 0:
            ws.column_dimensions[col_letter].width = max_length + 2

excel_lock = Lock()  
def excel2pdf(file):
    with excel_lock:
        pythoncom.CoInitialize()

        excel_app = None
        workbook = None

        try:
            file_path = os.path.abspath(file)

            if not os.path.exists(file_path):
                raise Exception(f"Excel file not found at: {file_path}")

            # Create a TRUE Excel Application instance
            excel_app = win32com.client.DispatchEx("Excel.Application")

            # Set properties on Application object
            excel_app.Application.DisplayAlerts = False
            excel_app.Visible = False

            workbook = excel_app.Workbooks.Open(file_path)

            output = os.path.splitext(file_path)[0] + ".pdf"

            workbook.ExportAsFixedFormat(0, output)

            workbook.Close(False)
            excel_app.Quit()

            os.remove(file_path)

            return FileResponse(
                output,
                media_type="application/pdf",
                filename=os.path.basename(output)
            )

        finally:
            try:
                if workbook:
                    workbook.Close(False)
                if excel_app:
                    excel_app.Quit()
            except:
                pass

            pythoncom.CoUninitialize()